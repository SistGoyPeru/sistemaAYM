import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
def exportar_proveedor_excel(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    scorecards = proveedor.scorecards.all().order_by('-fecha')
    incidentes = proveedor.incidentes.all().order_by('-fecha')
    ordenes = proveedor.ordenes_compra.all().order_by('-fecha')
    facturas = proveedor.facturas.all().order_by('-fecha')
    devoluciones = proveedor.devoluciones.all().order_by('-fecha')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proveedor"

    ws.append(["Reporte de Proveedor"])
    ws.append(["Razón Social", proveedor.razon_social])
    ws.append(["Nombre Comercial", proveedor.nombre_comercial])
    ws.append(["ID Fiscal", proveedor.id_fiscal])
    ws.append(["Teléfono", proveedor.telefono])
    ws.append(["Email", proveedor.email])
    ws.append(["Estado", proveedor.get_estado_display()])
    ws.append([])

    ws.append(["Scorecard"])
    ws.append(["Fecha", "Calidad", "Puntualidad", "Cumplimiento"])
    for s in scorecards:
        ws.append([s.fecha, s.calidad, s.puntualidad, s.cumplimiento])
    ws.append([])

# --- Exportar PDF ---
from django.shortcuts import get_object_or_404
def exportar_proveedor_pdf(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    scorecards = proveedor.scorecards.all().order_by('-fecha')
    incidentes = proveedor.incidentes.all().order_by('-fecha')
    ordenes = proveedor.ordenes_compra.all().order_by('-fecha')
    facturas = proveedor.facturas.all().order_by('-fecha')
    devoluciones = proveedor.devoluciones.all().order_by('-fecha')
    avg_calidad = round(sum(s.calidad for s in scorecards) / scorecards.count(), 2) if scorecards.exists() else None
    avg_puntualidad = round(sum(s.puntualidad for s in scorecards) / scorecards.count(), 2) if scorecards.exists() else None
    avg_cumplimiento = round(sum(s.cumplimiento for s in scorecards) / scorecards.count(), 2) if scorecards.exists() else None
    template = get_template('proveedores/detalle_pdf.html')
    html = template.render({
        'proveedor': proveedor,
        'scorecards': scorecards,
        'incidentes': incidentes,
        'avg_calidad': avg_calidad,
        'avg_puntualidad': avg_puntualidad,
        'avg_cumplimiento': avg_cumplimiento,
        'ordenes': ordenes,
        'facturas': facturas,
        'devoluciones': devoluciones,
        'pdf': True,
    })
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="proveedor_{proveedor.pk}.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

    ws.append(["Incidentes"])
    ws.append(["Fecha", "Descripción"])
    for i in incidentes:
        ws.append([i.fecha, i.descripcion])
    ws.append([])

    ws.append(["Órdenes de Compra"])
    ws.append(["Código", "Estado", "Fecha"])
    for o in ordenes:
        ws.append([o.codigo, o.get_estado_display(), o.fecha])
    ws.append([])

    ws.append(["Facturas"])
    ws.append(["Número", "Estado", "Fecha", "Monto"])
    for f in facturas:
        ws.append([f.numero, f.get_estado_display(), f.fecha, f.monto])
    ws.append([])

    ws.append(["Devoluciones"])
    ws.append(["Fecha", "Motivo", "Nota de Crédito"])
    for d in devoluciones:
        ws.append([d.fecha, d.motivo, d.nota_credito])

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="proveedor_{proveedor.pk}.xlsx"'
    wb.save(response)
    return response
def detalle_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    scorecards = proveedor.scorecards.all().order_by('-fecha')
    incidentes = proveedor.incidentes.all().order_by('-fecha')
    # Calcular promedios
    if scorecards.exists():
        avg_calidad = round(sum(s.calidad for s in scorecards) / scorecards.count(), 2)
        avg_puntualidad = round(sum(s.puntualidad for s in scorecards) / scorecards.count(), 2)
        avg_cumplimiento = round(sum(s.cumplimiento for s in scorecards) / scorecards.count(), 2)
    else:
        avg_calidad = avg_puntualidad = avg_cumplimiento = None
    ordenes = proveedor.ordenes_compra.all().order_by('-fecha_emision')
    facturas = proveedor.facturas.all().order_by('-fecha')
    devoluciones = proveedor.devoluciones.all().order_by('-fecha')
    return render(request, 'proveedores/detalle.html', {
        'proveedor': proveedor,
        'scorecards': scorecards,
        'incidentes': incidentes,
        'avg_calidad': avg_calidad,
        'avg_puntualidad': avg_puntualidad,
        'avg_cumplimiento': avg_cumplimiento,
        'ordenes': ordenes,
        'facturas': facturas,
        'devoluciones': devoluciones,
    })
from .models import ProductoProveedor
def catalogo_productos(request):
    productos = ProductoProveedor.objects.all()
    return render(request, 'proveedores/catalogo.html', {'productos': productos})

from django.shortcuts import render, get_object_or_404, redirect
from .models import Proveedor
from .forms import ProveedorForm, DatosBancariosFormSet, CondicionPagoFormSet, DocumentoLegalFormSet, MonedaFormSet, ProductoProveedorFormSet, IncidenteProveedorFormSet

def lista_proveedores(request):
    from datetime import date, timedelta
    proveedores = Proveedor.objects.all()
    q = request.GET.get('q', '').strip()
    estado_sel = request.GET.get('estado', '')
    if q:
        proveedores = proveedores.filter(
            Q(razon_social__icontains=q) |
            Q(nombre_comercial__icontains=q) |
            Q(id_fiscal__icontains=q) |
            Q(email__icontains=q)
        )
    if estado_sel:
        proveedores = proveedores.filter(estado=estado_sel)
    total = proveedores.count()
    activos = proveedores.filter(estado="activo").count()
    en_prueba = proveedores.filter(estado="en_prueba").count()
    bloqueados = proveedores.filter(estado="bloqueado").count()
    # Exportar a Excel
    if 'exportar' in request.GET:
        import openpyxl
        from django.http import HttpResponse
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proveedores"
        ws.append(["Razón Social", "Nombre Comercial", "ID Fiscal", "Categoría", "Teléfono", "Email", "Estado"])
        for p in proveedores:
            ws.append([
                p.razon_social,
                p.nombre_comercial,
                p.id_fiscal,
                p.categoria,
                p.telefono,
                p.email,
                p.get_estado_display(),
            ])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=proveedores.xlsx'
        wb.save(response)
        return response
    if request.GET.get('dashboard') == '1':
        # Promedios de scorecard por proveedor
        labels = []
        calidad = []
        puntualidad = []
        cumplimiento = []
        for p in proveedores:
            sc = p.scorecards.all()
            labels.append(p.razon_social)
            if sc.exists():
                calidad.append(round(sum(s.calidad for s in sc) / sc.count(), 2))
                puntualidad.append(round(sum(s.puntualidad for s in sc) / sc.count(), 2))
                cumplimiento.append(round(sum(s.cumplimiento for s in sc) / sc.count(), 2))
            else:
                calidad.append(0)
                puntualidad.append(0)
                cumplimiento.append(0)
        from django.http import JsonResponse
        return JsonResponse({
            'labels': labels,
            'calidad': calidad,
            'puntualidad': puntualidad,
            'cumplimiento': cumplimiento,
        })
    # Facturas vencidas: estado pendiente y fecha < hoy
    facturas_vencidas = []
    for p in proveedores:
        for f in p.facturas.filter(estado='pendiente', fecha__lt=date.today()):
            facturas_vencidas.append(f)
    # Incidentes críticos: incidentes en los últimos 7 días con palabra clave "crítico"
    incidentes_criticos = []
    hace_7dias = date.today() - timedelta(days=7)
    for p in proveedores:
        for i in p.incidentes.filter(fecha__gte=hace_7dias):
            if 'criti' in i.descripcion.lower():
                incidentes_criticos.append(i)
    estrellas = [1, 2, 3, 4, 5]
    return render(request, 'proveedores/lista.html', {
        'proveedores': proveedores,
        'total': total,
        'activos': activos,
        'en_prueba': en_prueba,
        'bloqueados': bloqueados,
        'facturas_vencidas': facturas_vencidas,
        'incidentes_criticos': incidentes_criticos,
        'q': q,
        'estado_sel': estado_sel,
        'estrellas': estrellas,
    })

def nuevo_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST, request.FILES)
        proveedor = None
        if form.is_valid():
            proveedor = form.save(commit=False)
        bancos_formset = DatosBancariosFormSet(request.POST, request.FILES, instance=proveedor)
        pago_formset = CondicionPagoFormSet(request.POST, instance=proveedor)
        doc_formset = DocumentoLegalFormSet(request.POST, request.FILES, instance=proveedor)
        moneda_formset = MonedaFormSet(request.POST, instance=proveedor)
        productos_formset = ProductoProveedorFormSet(request.POST, instance=proveedor)
        incidente_formset = IncidenteProveedorFormSet(request.POST, instance=proveedor)
        if form.is_valid() and bancos_formset.is_valid() and pago_formset.is_valid() and doc_formset.is_valid() and moneda_formset.is_valid() and productos_formset.is_valid() and incidente_formset.is_valid():
            proveedor.save()
            bancos_formset.instance = proveedor
            pago_formset.instance = proveedor
            doc_formset.instance = proveedor
            moneda_formset.instance = proveedor
            productos_formset.instance = proveedor
            incidente_formset.instance = proveedor
            bancos_formset.save()
            pago_formset.save()
            doc_formset.save()
            moneda_formset.save()
            productos_formset.save()
            incidente_formset.save()
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm()
        bancos_formset = DatosBancariosFormSet(instance=None)
        pago_formset = CondicionPagoFormSet(instance=None)
        doc_formset = DocumentoLegalFormSet(instance=None)
        moneda_formset = MonedaFormSet(instance=None)
        productos_formset = ProductoProveedorFormSet(instance=None)
        incidente_formset = IncidenteProveedorFormSet(instance=None)
    return render(request, 'proveedores/form.html', {
        'form': form,
        'bancos_formset': bancos_formset,
        'pago_formset': pago_formset,
        'doc_formset': doc_formset,
        'moneda_formset': moneda_formset,
        'productos_formset': productos_formset,
        'incidente_formset': incidente_formset,
    })


def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, request.FILES, instance=proveedor)
        bancos_formset = DatosBancariosFormSet(request.POST, request.FILES, instance=proveedor)
        pago_formset = CondicionPagoFormSet(request.POST, instance=proveedor)
        doc_formset = DocumentoLegalFormSet(request.POST, request.FILES, instance=proveedor)
        moneda_formset = MonedaFormSet(request.POST, instance=proveedor)
        productos_formset = ProductoProveedorFormSet(request.POST, instance=proveedor)
        incidente_formset = IncidenteProveedorFormSet(request.POST, instance=proveedor)
        if form.is_valid() and bancos_formset.is_valid() and pago_formset.is_valid() and doc_formset.is_valid() and moneda_formset.is_valid() and productos_formset.is_valid() and incidente_formset.is_valid():
            form.save()
            bancos_formset.save()
            pago_formset.save()
            doc_formset.save()
            moneda_formset.save()
            productos_formset.save()
            incidente_formset.save()
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
        bancos_formset = DatosBancariosFormSet(instance=proveedor)
        pago_formset = CondicionPagoFormSet(instance=proveedor)
        doc_formset = DocumentoLegalFormSet(instance=proveedor)
        moneda_formset = MonedaFormSet(instance=proveedor)
        productos_formset = ProductoProveedorFormSet(instance=proveedor)
        incidente_formset = IncidenteProveedorFormSet(instance=proveedor)
    return render(request, 'proveedores/form.html', {
        'form': form,
        'proveedor': proveedor,
        'bancos_formset': bancos_formset,
        'pago_formset': pago_formset,
        'doc_formset': doc_formset,
        'moneda_formset': moneda_formset,
        'productos_formset': productos_formset,
        'incidente_formset': incidente_formset,
    })

def eliminar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        proveedor.delete()
        return redirect('lista_proveedores')
    return render(request, 'proveedores/eliminar.html', {'proveedor': proveedor})
